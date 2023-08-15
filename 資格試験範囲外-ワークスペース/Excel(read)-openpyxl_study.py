import openpyxl

wb = openpyxl.load_workbook("Excel-openpyxl_study.xlsx")  # ワークブックオブジェクトを取得

# シート一覧の取得
print(wb.sheetnames)

ws = wb[wb.sheetnames[0]]  # ワークブックオブジェクトから、シート名を指定してワークシートオブジェクトを取得

# どちらかの方法でセルを取得できる
print(ws["B2"].value)  # ワークシートオブジェクトから、セルを指定して値を取得
print(ws.cell(row=3, column=3).value)  # ワークシートオブジェクトのcell()メソッドを使用して、rowとcolumnでセルを取得


print("------------")
for row in ws.rows:  # rowsはワークシートオブジェクトのプロパティで行を1行ずつ返すジェネレーター
    for cell in row:
        print(cell.value)  # A1は空白のため、Noneが返る
