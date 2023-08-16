import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# 1.Excel用オブジェクトを作成
wb = openpyxl.Workbook()

# 2.シートを作成
ws = wb.create_sheet(title="New Sheet", index=0)  # index=0を指定してシートを一番左に挿入する

# 3.セルに書き込み
ws["A1"] = 100  # セルA1に100を入力

# フォントを指定することもできる
font = Font(
    name="Arial",  # フォント名
    sz=13,  # 文字サイズ
    bold=True,  # 太字
    underline=None,  # 下線を引く
    strike=True,  # 打ち消し線
    vertAlign="subscript",  # 文字の配置（垂直方向）。subscript、baseline、superscriptから選択
    color="FF000000",  # 色の指定
)
ws["A2"].font = font
ws["A2"] = 200

fill = PatternFill(
    fill_type="solid",
    fgColor="FF0000",
    bgColor="FF0000",
)

ws["A3"].fill = fill

# 4.保存
wb.save(filename="new_book.xlsx")
